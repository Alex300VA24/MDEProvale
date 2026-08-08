#!/usr/bin/env python3
"""
generate_seeders.py
--------------------
Lee el Excel "BASE DE DATOS PROVALE" (padrones de beneficiarios de varios
Clubes de Madres apilados uno debajo del otro en una sola hoja) y genera
5 seeders de Laravel:

    - AssociationSeeder.php
    - PeopleSeeder.php
    - PartnerSeeder.php
    - BeneficiarieSeeder.php
    - BeneficiarieHistorySeeder.php

IMPORTANTE - LEE ESTO ANTES DE CORRER EN PRODUCCION
=====================================================
El Excel NO trae los IDs reales de tu base de datos para varias columnas
(relationship_id, type_benefit_id, place_sector_id, resolution_id,
type_premises_id, reason_disqualification_id, state_id, position_id).
Esos catálogos viven en tu base de datos (Laravel), no en el Excel.

Por eso este script trae, al inicio, una sección "CONFIGURACION" con
diccionarios que TU DEBES completar con los IDs reales de tus tablas
`relationships`, `type_benefits`, `place_sectors`, etc. Mientras no los
completes, el script usa valores por defecto (documentados abajo) y
además imprime un reporte de advertencias (unmapped_report.txt) con
todos los códigos que no tienen mapeo, para que sepas exactamente qué
revisar antes de correr los seeders contra la base real.

Uso:
    python3 generate_seeders.py /ruta/al/excel.xlsx /ruta/salida

Requiere: openpyxl (pip install openpyxl --break-system-packages)
"""

import sys
import os
import re
import json
import datetime
from collections import OrderedDict, Counter

import openpyxl


# ============================================================================
# CONFIGURACION - EDITA ESTA SECCION CON LOS IDs REALES DE TU BASE DE DATOS
# ============================================================================

# Todas las asociaciones de este archivo son "Club de Madres" (se detecta
# por el texto "CLUB DE MADRES" en el titulo de cada bloque). Si en algun
# futuro Excel aparecen "Comite de Vaso de Leche" (CVL) or "OSB", agrega
# aqui el patron -> codigo de empresa.
COMPANY_NAME_PATTERNS = [
    (r"CLUB DE MADRES", "CDM"),
    (r"VASO DE LECHE", "CVL"),
    (r"OLLA COM", "OSB"),
]
DEFAULT_COMPANY_NAME = "CDM"

# parentesco (columna "parentesco" del Excel) -> title de tu tabla
# `relationships` (Socio / Hijos / Apoderado / Tutelado, segun tu
# RelationshipSeeder). El Excel no trae una leyenda de codigos, asi que
# esta tabla se infirio revisando manualmente muestras de cada codigo
# (comparando apellidos socia/beneficiario y edad). Es una inferencia
# razonable, NO una certeza -- revisala con tu equipo antes de correr en
# produccion:
#   '01'/'1' -> misma persona que la socia (LAC)                -> Socio
#   '03'/'3' -> apellido del beneficiario coincide con el de la
#               socia (patron tipico de hijo/a)                  -> Hijos
#   '02','05','08' -> beneficiario adulto mayor (65+), con
#               apellido parcialmente compartido con la socia
#               (ej. su madre/padre) que ella representa          -> Apoderado
#   '04','06','11' -> beneficiario nino/a pero SIN apellido en
#               comun con la socia (no séria su hijo/a biologico) -> Tutelado
PARENTESCO_TO_RELATIONSHIP_TITLE = {
    "01": "Socio",
    "1": "Socio",
    "03": "Hijos",
    "3": "Hijos",
    "02": "Apoderado",
    "05": "Apoderado",
    "08": "Apoderado",
    "04": "Tutelado",
    "06": "Tutelado",
    "11": "Tutelado",
}
DEFAULT_RELATIONSHIP_TITLE = None  # usado si aparece un codigo no listado arriba

# tipo_benef (columna "tipo_benef": '', 'LAC', 'GES', 'DIS') -> abbreviation
# real de tu tabla `type_benefits` (de tu TypeBenefitSeeder: NI0, NI7, ADU,
# GES, LAC, TBC, DIS). Cuando tipo_benef viene vacio, se infiere NI0/NI7/ADU
# a partir de la edad (0-6, 7-13, 65+ respectivamente -- asi se distribuyen
# limpiamente todas las filas vacias de este archivo, sin casos 14-64).
TIPO_BENEF_TO_TYPE_BENEFIT_ABBR = {
    "LAC": "LAC",
    "GES": "GES",
    "DIS": "DIS",
}


def infer_type_benefit_abbr(tipo_benef, edad):
    tb = (tipo_benef or "").strip().upper()
    if tb in TIPO_BENEF_TO_TYPE_BENEFIT_ABBR:
        return TIPO_BENEF_TO_TYPE_BENEFIT_ABBR[tb]
    if edad is None:
        return None
    if edad <= 6:
        return "NI0"
    if edad <= 13:
        return "NI7"
    if edad >= 65:
        return "ADU"
    return None  # rango 14-64 sin flag explicito: no encontrado en este archivo

# Codigo de asociacion (columna "ncodigo", ej. '005') -> place_sector_id real.
# El Excel NO trae el sector/distrito como ID, solo la direccion en texto.
# Dejalo vacio y complementa manualmente, o agrega aqui los que conozcas:
#   "005": 1,
#   "015": 3,
ASSOCIATION_CODE_TO_PLACE_SECTOR_ID = {}
DEFAULT_PLACE_SECTOR_ID = None

# Mismo problema para resolution_id / type_premises_id de `associations`:
# no vienen en el Excel. Si tienes un archivo aparte con esa info,
# mapealo aqui por codigo de asociacion.
ASSOCIATION_CODE_TO_RESOLUTION_ID = {}
DEFAULT_RESOLUTION_ID = None

# type_premises SI tiene catalogo conocido (TypePremisesSeeder: Propio,
# Provisional, Municipalidad), pero el Excel no trae esta info por
# asociacion (no hay ninguna columna ni texto al respecto). Completa aqui
# el TITLE correspondiente por codigo de asociacion, ej.:
#   "005": "Propio",
#   "015": "Municipalidad",
ASSOCIATION_CODE_TO_TYPE_PREMISES_TITLE = {}

# Valores por defecto para columnas de estado que si suelen tener un
# valor "activo" estandar (ajusta si tu catalogo usa otro id).
DEFAULT_STATE_ID = 1
DEFAULT_REASON_DISQUALIFICATION_ID = None
DEFAULT_POSITION_ID = None

# El reporte no trae "date_begin" de beneficiary_histories (esa es la fecha
# de nacimiento del beneficiario, ya usada como `birthdate` en People).
# Como aproximacion se usa el inicio del periodo que indica el propio
# encabezado del Excel ("AÑO 2026 - I SEMESTRE"). Ajusta si corresponde.
BENEFICIARY_HISTORY_DATE_BEGIN = "2026-01-01"

# Cuantos parametros (columnas x filas) como maximo por chunk de INSERT.
# Los seeders de ejemplo usan ~2000 parametros por chunk.
MAX_PARAMS_PER_CHUNK = 2000


# ============================================================================
# PARSEO DEL EXCEL
# ============================================================================

def col_letter_to_index(_):
    raise NotImplementedError


def normalize_dni(value):
    if value is None:
        return None
    if isinstance(value, float):
        value = int(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def normalize_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def normalize_date(value):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    return s if s else None


def normalize_gender(value):
    if value is None:
        return None
    s = str(value).strip().upper()
    if s in ("M", "F"):
        return s
    return None


def guess_company_name(title):
    for pattern, code in COMPANY_NAME_PATTERNS:
        if re.search(pattern, title, re.IGNORECASE):
            return code
    return DEFAULT_COMPANY_NAME


def parse_workbook(path):
    """
    Devuelve una lista de "bloques", uno por asociacion/club encontrado en
    la hoja. Cada bloque trae los datos del encabezado (codigo, nombre,
    presidenta, direccion) y la lista de filas de datos (beneficiarios).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Detectar inicio de cada bloque: fila cuyo texto de columna B contiene
    # "PADRÓN" / "PADRON".
    starts = [
        i for i, r in enumerate(rows)
        if len(r) > 1 and r[1] and isinstance(r[1], str) and "PADR" in r[1].upper()
    ]
    if not starts:
        raise ValueError("No se encontraron bloques 'PADRÓN DE BENEFICIARIOS' en el archivo.")

    blocks = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] if idx + 1 < len(starts) else len(rows)

        title_row = rows[start]
        title = normalize_str(title_row[1]) or ""
        # Codigo de asociacion viene en la misma fila del titulo, columna N (index 13/14)
        code = None
        for c in range(len(title_row)):
            if title_row[c] == "COD" and c + 1 < len(title_row):
                code = normalize_str(title_row[c + 1])
                break

        # nombre del club: texto despues de "CLUB DE MADRES:" (o similar)
        name_match = re.search(r":\s*(.+)$", title)
        club_name = name_match.group(1).strip() if name_match else title

        presidenta_row = rows[start + 3] if start + 3 < len(rows) else ()
        presidenta = None
        address = None
        for cell in presidenta_row:
            if isinstance(cell, str):
                if cell.strip().startswith("Presidenta"):
                    presidenta = cell.split(":", 1)[1].strip().rstrip(".") if ":" in cell else None
                elif cell.strip().startswith("Direcci"):
                    address = cell.split(":", 1)[1].strip() if ":" in cell else None

        # fila de encabezados de columnas ("ncodigo", ...) deberia estar en start+5
        header_row_idx = start + 5
        data_start = header_row_idx + 1

        data_rows = []
        for i in range(data_start, end):
            r = rows[i]
            if len(r) < 17:
                continue
            # Fila de datos valida: columna B (indice 1) es un numero de
            # secuencia y columna A (indice 0) tiene el codigo de asociacion.
            if r[0] is not None and isinstance(r[1], (int, float)):
                data_rows.append(r)

        blocks.append({
            "code": code,
            "club_name": club_name,
            "presidenta": presidenta,
            "address": address,
            "company_name": guess_company_name(title),
            "data_rows": data_rows,
        })

    return blocks


# ============================================================================
# NORMALIZACION A ENTIDADES (people / associations / partners / beneficiaries)
# ============================================================================

def build_entities(blocks):
    associations = []          # una por bloque
    people = OrderedDict()     # dni -> dict
    partners = OrderedDict()   # (dni_socia, code) -> dict
    beneficiaries = OrderedDict()  # dni_beneficiario -> dict

    unmapped_parentesco = Counter()
    unmapped_tipo_benef = Counter()
    rows_missing_dni = 0

    for block in blocks:
        code = block["code"]
        associations.append({
            "code": code,
            "name": block["club_name"],
            "company_name": block["company_name"],
            "address": block["address"],
            "presidenta": block["presidenta"],
        })

        for r in block["data_rows"]:
            socia_ape_pat = normalize_str(r[2])
            socia_ape_mat = normalize_str(r[3])
            socia_nombres = normalize_str(r[4])
            socia_dir = normalize_str(r[5])
            socia_dni = normalize_dni(r[6])
            socia_fn = normalize_date(r[7])

            benef_ape_pat = normalize_str(r[8])
            benef_ape_mat = normalize_str(r[9])
            benef_nombres = normalize_str(r[10])
            benef_dni = normalize_dni(r[11])

            parentesco = normalize_str(r[12]) or ""
            tipo_benef = r[13] if r[13] is not None else ""
            tipo_benef = str(tipo_benef).strip()
            benef_fecnac = normalize_date(r[14])
            benef_sexo = normalize_gender(r[16])
            benef_edad = r[15] if isinstance(r[15], (int, float)) else None

            if not socia_dni or not benef_dni:
                rows_missing_dni += 1
                continue

            # --- people: socia ---
            if socia_dni not in people:
                people[socia_dni] = {
                    "dni": socia_dni,
                    "father_lastname": socia_ape_pat,
                    "mother_lastname": socia_ape_mat,
                    "names": socia_nombres,
                    "gender": "F",  # todas las socias de "Club de Madres" son mujeres
                    "birthdate": socia_fn,
                    "address": socia_dir,
                    "association_code": code,
                }

            # --- people: beneficiario ---
            if benef_dni not in people:
                people[benef_dni] = {
                    "dni": benef_dni,
                    "father_lastname": benef_ape_pat,
                    "mother_lastname": benef_ape_mat,
                    "names": benef_nombres,
                    "gender": benef_sexo,
                    "birthdate": benef_fecnac,
                    # el Excel no trae direccion propia del beneficiario:
                    # se asume la misma direccion que su socia/apoderada.
                    "address": socia_dir,
                    "association_code": code,
                }

            # --- partners: una fila por socia unica dentro de su asociacion ---
            partner_key = (socia_dni, code)
            if partner_key not in partners:
                partners[partner_key] = {
                    "person_dni": socia_dni,
                    "association_code": code,
                    "observations": "Socia titular",
                }

            # --- beneficiaries: una fila por beneficiario unico ---
            relationship_title = PARENTESCO_TO_RELATIONSHIP_TITLE.get(parentesco, DEFAULT_RELATIONSHIP_TITLE)
            type_benefit_abbr = infer_type_benefit_abbr(tipo_benef, benef_edad)

            if relationship_title is None:
                unmapped_parentesco[parentesco] += 1
            if type_benefit_abbr is None:
                unmapped_tipo_benef[f"tipo_benef={tipo_benef!r} edad={benef_edad!r}"] += 1

            if benef_dni not in beneficiaries:
                beneficiaries[benef_dni] = {
                    "person_dni": benef_dni,
                    "partner_dni": socia_dni,
                    "association_code": code,
                    "parentesco": parentesco,
                    "relationship_title": relationship_title,
                    "tipo_benef": tipo_benef,
                    "type_benefit_abbr": type_benefit_abbr,
                    "birthdate": benef_fecnac,
                }

    warnings = {
        "unmapped_parentesco": dict(unmapped_parentesco),
        "unmapped_tipo_benef": dict(unmapped_tipo_benef),
        "rows_missing_dni": rows_missing_dni,
    }

    return associations, people, partners, beneficiaries, warnings


# ============================================================================
# HELPERS PARA GENERAR PHP
# ============================================================================

def php_escape(value):
    return value.replace("\\", "\\\\").replace("'", "\\'")


def php_str(value):
    if value is None:
        return "null"
    return "'" + php_escape(str(value)) + "'"


def php_raw(value):
    """Para valores numericos / expresiones ya validas en PHP (int, null, o codigo)."""
    if value is None:
        return "null"
    return str(value)


def chunk_by_params(records, num_cols, max_params=MAX_PARAMS_PER_CHUNK):
    per_chunk = max(1, max_params // num_cols)
    for i in range(0, len(records), per_chunk):
        yield records[i:i + per_chunk]


PHP_HEADER = """<?php

namespace Database\\Seeders;

use Illuminate\\Database\\Seeder;
use Illuminate\\Support\\Facades\\DB;

class {class_name} extends Seeder
{{
    public function run(): void
    {{
"""

PHP_FOOTER = """    }
}
"""


# ============================================================================
# GENERADORES POR SEEDER
# ============================================================================

def generate_association_seeder(associations):
    num_cols = 9  # code,name,company_name,address,resolution_id,state_id,place_sector_id,type_premises_id,+timestamps(2) -> ver abajo
    lines = []
    lines.append(PHP_HEADER.format(class_name="AssociationSeeder").rstrip("\n"))
    lines.append("        /**")
    lines.append("         * Asociaciones/Comités extraídos del Excel (padrón de beneficiarios).")
    lines.append("         * OJO: resolution_id, place_sector_id y type_premises_id NO vienen en el")
    lines.append("         * Excel; revisa la sección CONFIGURACION en generate_seeders.py.")
    lines.append("         */")
    lines.append("        $associations = [")
    for a in associations:
        resolution_id = ASSOCIATION_CODE_TO_RESOLUTION_ID.get(a["code"], DEFAULT_RESOLUTION_ID)
        place_sector_id = ASSOCIATION_CODE_TO_PLACE_SECTOR_ID.get(a["code"], DEFAULT_PLACE_SECTOR_ID)
        type_premises_title = ASSOCIATION_CODE_TO_TYPE_PREMISES_TITLE.get(a["code"])
        lines.append("            [")
        lines.append(f"                'code'             => {php_str(a['code'])},")
        lines.append(f"                'name'             => {php_str(a['name'])},")
        lines.append(f"                'company_name'     => {php_str(a['company_name'])},")
        lines.append(f"                'address'          => {php_str(a['address'])},")
        if a.get("presidenta"):
            lines.append(f"                #'president'       => {php_str(a['presidenta'])},")
        lines.append(f"                'resolution_id'    => {php_raw(resolution_id)},")
        lines.append(f"                'state_id'         => {php_raw(DEFAULT_STATE_ID)},")
        lines.append(f"                'place_sector_id'  => {php_raw(place_sector_id)},")
        if type_premises_title:
            lines.append(f"                'type_premises_id' => DB::table('type_premises')->where('title', {php_str(type_premises_title)})->value('id'),")
        else:
            lines.append(f"                'type_premises_id' => null, // TODO: definir tipo de local para code={a['code']!r} (Propio/Provisional/Municipalidad)")
        lines.append("                'created_at'       => now(),")
        lines.append("                'updated_at'       => now(),")
        lines.append("            ],")
    lines.append("        ];")
    lines.append("")
    lines.append("        DB::table('associations')->insert($associations);")
    lines.append(PHP_FOOTER)
    return "\n".join(lines)


def generate_people_seeder(people):
    records = list(people.values())
    num_cols = 10  # names, father_lastname, mother_lastname, dni, gender, telephone_number, phone_number, birthdate, address, place_sector_id (+created_at/updated_at = 12)
    num_cols = 12
    out = [PHP_HEADER.format(class_name="PeopleSeeder").rstrip("\n")]
    total_chunks = list(chunk_by_params(records, num_cols))
    for ci, chunk in enumerate(total_chunks, start=1):
        out.append(f"        // Chunk {ci} ({len(chunk)} registros — {num_cols} cols × {len(chunk)} = {num_cols * len(chunk)} params)")
        out.append("        DB::table('people')->insert([")
        for p in chunk:
            place_sector_id = ASSOCIATION_CODE_TO_PLACE_SECTOR_ID.get(p["association_code"], DEFAULT_PLACE_SECTOR_ID)
            out.append("            [")
            out.append(f"                'names'            => {php_str(p['names'])},")
            out.append(f"                'father_lastname'  => {php_str(p['father_lastname'])},")
            out.append(f"                'mother_lastname'  => {php_str(p['mother_lastname'])},")
            out.append(f"                'dni'              => {php_str(p['dni'])},")
            out.append(f"                'gender'           => {php_str(p['gender'])},")
            out.append("                'telephone_number' => null,")
            out.append("                'phone_number'     => null,")
            out.append(f"                'birthdate'        => {php_str(p['birthdate'])},")
            out.append(f"                'address'          => {php_str(p['address'])},")
            out.append(f"                'place_sector_id'  => {php_raw(place_sector_id)},")
            out.append("                'created_at'       => now(),")
            out.append("                'updated_at'       => now(),")
            out.append("            ],")
        out.append("        ]);")
        out.append("")
    out.append(PHP_FOOTER)
    return "\n".join(out)


def generate_partner_seeder(partners):
    records = list(partners.values())
    num_cols = 9
    out = [PHP_HEADER.format(class_name="PartnerSeeder").rstrip("\n")]
    total_chunks = list(chunk_by_params(records, num_cols))
    for ci, chunk in enumerate(total_chunks, start=1):
        out.append(f"        // Chunk {ci} ({len(chunk)} registros — {num_cols} cols × {len(chunk)} = {num_cols * len(chunk)} params)")
        out.append("        DB::table('partners')->insert([")
        for p in chunk:
            out.append("            [")
            out.append("                'date_begin'     => null,")
            out.append("                'date_end'       => null,")
            out.append(f"                'observations'   => {php_str(p['observations'])},")
            out.append(f"                'person_id'      => DB::table('people')->where('dni', {php_str(p['person_dni'])})->value('id'),")
            out.append(f"                'association_id' => DB::table('associations')->where('code', {php_str(p['association_code'])})->value('id'),")
            out.append(f"                'state_id'       => {php_raw(DEFAULT_STATE_ID)},")
            out.append(f"                'position_id'    => {php_raw(DEFAULT_POSITION_ID)},")
            out.append("                'created_at'     => now(),")
            out.append("                'updated_at'     => now(),")
            out.append("            ],")
        out.append("        ]);")
        out.append("")
    out.append(PHP_FOOTER)
    return "\n".join(out)


def generate_beneficiarie_seeder(beneficiaries):
    records = list(beneficiaries.values())
    num_cols = 5
    out = [PHP_HEADER.format(class_name="BeneficiarieSeeder").rstrip("\n")]
    total_chunks = list(chunk_by_params(records, num_cols))
    for ci, chunk in enumerate(total_chunks, start=1):
        out.append(f"        // Chunk {ci} ({len(chunk)} registros — {num_cols} cols × {len(chunk)} = {num_cols * len(chunk)} params)")
        out.append("        DB::table('beneficiaries')->insert([")
        for b in chunk:
            rel_title = b["relationship_title"]
            out.append("            [")
            out.append(f"                'person_id'       => DB::table('people')->where('dni', {php_str(b['person_dni'])})->value('id'),")
            out.append("                'partner_id'      => DB::table('partners')")
            out.append("                                        ->where('person_id',")
            out.append(f"                                            DB::table('people')->where('dni', {php_str(b['partner_dni'])})->value('id')")
            out.append("                                        )->value('id'),")
            if rel_title is None:
                out.append(f"                'relationship_id' => null, // TODO: mapear parentesco={b['parentesco']!r}")
            else:
                out.append(f"                'relationship_id' => DB::table('relationships')->where('title', {php_str(rel_title)})->value('id'),")
            out.append("                'created_at'      => now(),")
            out.append("                'updated_at'      => now(),")
            out.append("            ],")
        out.append("        ]);")
        out.append("")
    out.append(PHP_FOOTER)
    return "\n".join(out)


def generate_beneficiarie_history_seeder(beneficiaries):
    records = list(beneficiaries.values())
    num_cols = 11
    out = [PHP_HEADER.format(class_name="BeneficiarieHistorySeeder").rstrip("\n")]
    total_chunks = list(chunk_by_params(records, num_cols))
    for ci, chunk in enumerate(total_chunks, start=1):
        out.append(f"        // Chunk {ci} ({len(chunk)} registros — {num_cols} cols × {len(chunk)} = {num_cols * len(chunk)} params)")
        out.append("        DB::table('beneficiary_histories')->insert([")
        for b in chunk:
            type_benefit_abbr = b["type_benefit_abbr"]
            out.append("            [")
            out.append("                'weight'                     => 0.0,")
            out.append("                'height'                     => 0.0,")
            out.append("                'hmg'                        => 0.0,")
            out.append(f"                'date_begin'                 => {php_str(BENEFICIARY_HISTORY_DATE_BEGIN)},")
            out.append("                'date_end'                   => null,")
            if type_benefit_abbr is None:
                out.append(f"                'type_benefit_id'            => null, // TODO: mapear tipo_benef={b['tipo_benef']!r}")
            else:
                out.append(f"                'type_benefit_id'            => DB::table('type_benefits')->where('abbreviation', {php_str(type_benefit_abbr)})->value('id'),")
            out.append("                'beneficiary_id'             => DB::table('beneficiaries')")
            out.append("                                                    ->where('person_id',")
            out.append(f"                                                        DB::table('people')->where('dni', {php_str(b['person_dni'])})->value('id')")
            out.append("                                                    )->value('id'),")
            out.append(f"                'state_id'                   => {php_raw(DEFAULT_STATE_ID)},")
            out.append(f"                'reason_disqualification_id' => {php_raw(DEFAULT_REASON_DISQUALIFICATION_ID)},")
            out.append("                'created_at'                 => now(),")
            out.append("                'updated_at'                 => now(),")
            out.append("            ],")
        out.append("        ]);")
        out.append("")
    out.append(PHP_FOOTER)
    return "\n".join(out)


# ============================================================================
# MAIN
# ============================================================================

def main():
    if len(sys.argv) < 3:
        print("Uso: python3 generate_seeders.py <excel.xlsx> <carpeta_salida>")
        sys.exit(1)

    excel_path = sys.argv[1]
    out_dir = sys.argv[2]
    os.makedirs(out_dir, exist_ok=True)

    print(f"Leyendo {excel_path} ...")
    blocks = parse_workbook(excel_path)
    print(f"  {len(blocks)} asociaciones/clubes detectados.")

    associations, people, partners, beneficiaries, warnings = build_entities(blocks)
    print(f"  {len(people)} personas unicas (socias + beneficiarios).")
    print(f"  {len(partners)} socias (partners).")
    print(f"  {len(beneficiaries)} beneficiarios unicos.")

    files = {
        "AssociationSeeder.php": generate_association_seeder(associations),
        "PeopleSeeder.php": generate_people_seeder(people),
        "PartnerSeeder.php": generate_partner_seeder(partners),
        "BeneficiarieSeeder.php": generate_beneficiarie_seeder(beneficiaries),
        "BeneficiarieHistorySeeder.php": generate_beneficiarie_history_seeder(beneficiaries),
    }

    for filename, content in files.items():
        path = os.path.join(out_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"  Generado: {path}")

    # Reporte de advertencias / mapeos pendientes
    report_path = os.path.join(out_dir, "unmapped_report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("REPORTE DE MAPEOS PENDIENTES\n")
        f.write("=============================\n\n")
        f.write("Filas con DNI faltante (ignoradas): {}\n\n".format(warnings["rows_missing_dni"]))
        f.write("Codigos de 'parentesco' sin mapeo a relationship title (columna configurable\n")
        f.write("en PARENTESCO_TO_RELATIONSHIP_TITLE):\n")
        for code, count in sorted(warnings["unmapped_parentesco"].items(), key=lambda x: -x[1]):
            f.write(f"  - {code!r}: {count} beneficiarios\n")
        f.write("\n")
        f.write("Casos sin abreviatura de type_benefit inferida (revisar infer_type_benefit_abbr):\n")
        for code, count in sorted(warnings["unmapped_tipo_benef"].items(), key=lambda x: -x[1]):
            f.write(f"  - {code}: {count} beneficiarios\n")
        f.write("\n")
        f.write("Recuerda tambien completar (si aplica):\n")
        f.write("  - ASSOCIATION_CODE_TO_PLACE_SECTOR_ID (no viene en el Excel)\n")
        f.write("  - ASSOCIATION_CODE_TO_RESOLUTION_ID (no viene en el Excel)\n")
        f.write("  - ASSOCIATION_CODE_TO_TYPE_PREMISES_TITLE (no viene en el Excel; usa\n")
        f.write("    'Propio' / 'Provisional' / 'Municipalidad' segun TypePremisesSeeder)\n")
    print(f"  Reporte de mapeos pendientes: {report_path}")

    print("\nListo. Revisa unmapped_report.txt antes de correr los seeders en produccion.")


if __name__ == "__main__":
    main()